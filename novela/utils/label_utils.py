# coding=utf-8
# @Author: 莫冉
# @Date: 2021-02-16
import os
import jieba
import numpy as np
from typing import Set, List, Optional, Tuple, Dict, Any

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.merge import MergedCellRange
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, NamedStyle
from openpyxl.styles import Side, Border, Alignment, PatternFill

from novela import logger
from novela.label import BaseLabel, Label
from novela.text import WordVectorSimilarity, HowNetSimilarity, CilinSimilarity, SentVectorSimilarity


def cut_and_remove_stopwords(sentence: str, stopwords: Set[str]):
    """对句子分词并去除停用词"""
    if stopwords is not None:
        result = [word for word in jieba.cut(sentence) if word not in stopwords]
    else:
        result = jieba.lcut(sentence)
    return result


def get_mean_sim(sim_matrix: np.ndarray, masks: np.ndarray):
    """得到每一个标签对于单词表的平均相似度"""
    sim_sum = np.sum(sim_matrix, axis=1)    # 得到每一个标签对于词表中所有单词相似度的和
    valid_sum = np.sum(masks, axis=1)       # 得到每一个标签对应的有效词表数
    sim_mean = sim_sum / (valid_sum + 1e-5)
    return sim_mean


def get_max_sim(sim_matrix: np.ndarray, masks: np.ndarray = None):
    """得到每一个标签单词对于词表中单词的最大相似度"""
    sim_max = np.max(sim_matrix, axis=1)
    return sim_max


def similarity_between_base_label_and_sents(sent_words: List[str],
                                            sent_strings: List[str],
                                            base_label: BaseLabel,
                                            sim_word2vector: WordVectorSimilarity,
                                            sim_hownet: HowNetSimilarity,
                                            sim_cilin: CilinSimilarity,
                                            sim_sent2vector: SentVectorSimilarity,
                                            return_counts: int = 1):
    """计算baselabel中的标签以及各个标签描述和文本之间的相似度"""
    # 这个表示所有枚举类的英文名
    base_enum_names = base_label.enum_names
    # ---- 首先获取当前base_label标签中所有标签对应的单词名称 -------
    base_label_names = base_label.display_names
    if len(base_enum_names) <= 0:
        return None, None
    # 所有的相似度都已经归一化到0-1之间了
    # sim 的维度为 [label_num, words_num]
    # mask的维度为 [label_num, words_num]，1表示是有效值，0表示无效值
    # 计算word2vector相似度
    w2v_sim, w2v_mask = sim_word2vector.wordlist_sim(base_label_names,
                                                     sent_words)
    # # 计算cilin相似度
    # cilin_sim, cilin_mask = sim_cilin.wordlist_sim(base_label_names,
    #                                                sent_words)
    # 计算hownet相似度
    # hownet_sim, hownet_mask = sim_hownet.wordlist_sim(base_label_names,
    #                                                   sent_words)
    # 得到每一种单词相似度的均值和最大值
    w2v_sim_mean = get_mean_sim(w2v_sim, w2v_mask)
    w2v_sim_max = get_max_sim(w2v_sim, w2v_mask)
    # cilin_sim_mean = get_mean_sim(cilin_sim, cilin_mask)
    # cilin_sim_max = get_max_sim(cilin_sim, cilin_mask)
    # hownet_sim_mean = get_mean_sim(hownet_sim, hownet_mask)
    # hownet_sim_max = get_max_sim(hownet_sim, hownet_mask)

    word_sim_mean = w2v_sim_mean
    word_sim_max = w2v_sim_max
    word_sim = (word_sim_mean + word_sim_max) / 2


    # ---- 接着考虑该标签各个类别是否存在描述，计算描述和各个句子之间的相似度 ---------
    base_label_descriptions = base_label.descriptions
    if base_label_descriptions and len(base_label_descriptions) == len(base_label_names):
        # 句子的相似度也已经归一化到0-1之间，形状为 [label_num, sentence_num]
        s2v_sim, s2v_mask = sim_sent2vector.sentlist_sim(base_label_descriptions,
                                                         sent_strings)
        s2v_sim_mean = get_mean_sim(s2v_sim, s2v_mask)
        s2v_sim_max = get_max_sim(s2v_sim, s2v_mask)
        sent_sim = (s2v_sim_mean + s2v_sim_max) / 2
    else:
        sent_sim = None

    # --- 计算出所有相似度之后进行融合 ---
    if sent_sim is None:
        sim = word_sim
    else:
        sim = (word_sim + sent_sim) / 2

    # 得到最大相似度的索引以及对应的标签
    if return_counts == 1:
        max_index = np.argmax(sim)
        value = base_label_names[max_index]
        enum_name = base_enum_names[max_index]
        return enum_name, value
    elif return_counts > 1:
        indexes = np.argsort(sim)[::-1]   # 从大到小
        max_indexes = indexes[:return_counts]
        values = np.array(base_label_names)[max_indexes]
        enum_names = np.array(base_enum_names)[max_indexes]
        return enum_names.tolist(), values.tolist()
    else:
        raise ValueError(f"`return_count` must be greater equal than 1, but get `{return_counts}`.")


# --------------------------------------------------------------------------


def _get_column_max_chars(column: Tuple[Cell, ...]):
    """
    获取当前列列表中所有单元格中最大字符数
    """
    max_col_chars = 0
    for col in column:
        cur_col_chars = len(str(col.value))
        if max_col_chars < cur_col_chars:
            max_col_chars = cur_col_chars
    return max_col_chars


def _set_column_width(worksheet: Worksheet, col: int, max_col_chars: int, max_width: int = 100, add_width: int = 10):
    """
    设置某一列的单元格宽度
    """
    col_letter = get_column_letter(col)
    if max_col_chars > max_width:
        worksheet.column_dimensions[col_letter].width = 100
    elif max_col_chars > 2:
        worksheet.column_dimensions[col_letter].width = max_col_chars + add_width


def _adjust_cell_height(worksheet: Worksheet,
                        fixed_height: int = 20,
                        active_row: Optional[int] = None):
    """
    将某行或者整张表的单元格调整到固定高度
    """
    if active_row is not None:
        worksheet.row_dimensions[active_row].height = fixed_height
    else:
        for i in range(1, worksheet.max_row + 1):
            worksheet.row_dimensions[i].height = fixed_height


def _adjust_cell_width(worksheet: Worksheet,
                       max_width: int = 150,
                       add_width: int = 10,
                       active_column: Optional[int] = None):
    """
    将某列或者整张表格的单元格自适应调整到一定的宽度
    """
    columns = worksheet.columns
    if active_column is not None:
        column = columns[active_column-1]
        max_col_chars = _get_column_max_chars(column)
        _set_column_width(worksheet,
                          col=active_column,
                          max_col_chars=max_col_chars,
                          max_width=max_width,
                          add_width=add_width)
    else:
        for i, column in enumerate(columns):
            max_col_chars = _get_column_max_chars(column)
            _set_column_width(worksheet,
                              col=i+1,
                              max_col_chars=max_col_chars,
                              max_width=max_width,
                              add_width=add_width)


def _adjust_cell_width_and_height(worksheet: Worksheet,
                                  max_width: int = 150,
                                  add_width: int = 10,
                                  fixed_height: int = 20,
                                  active_row: Optional[int] = None,
                                  active_column: Optional[int] = None):
    """
    调节某一行或者表格中所有行列的高度和宽度
    :param worksheet: 当前操作的Worksheet对象
    :param max_width: int型，表示单元格允许的最大宽度
    :param add_width: int型，表示对于某个单元格宽度需要基于最大字符数的增加量
    :param fixed_height: int型，表示某个单元格允许的高度
    :param active_row: int型（可选），如果需要修改某一行的高度则指定（从1开始）
    :param active_column: int型（可选），如果需要修改某一列的宽度则指定（从1开始）
    :return:
    """
    # 首先调节行高度
    _adjust_cell_height(worksheet,
                        fixed_height=fixed_height,
                        active_row=active_row)
    # 接着调节列宽度
    _adjust_cell_width(worksheet,
                       max_width=max_width,
                       add_width=add_width,
                       active_column=active_column)


def _register_header_style(workbook: Workbook):
    """注册header的风格"""
    # 字体
    header_font = Font(name="微软雅黑", size=11, b=True)
    # 边框
    line_m = Side(style="medium", color="000000")  # 粗边框
    border_m = Border(top=line_m, bottom=line_m, left=line_m, right=line_m)
    # 对齐
    alignment = Alignment(horizontal="center", vertical="center")
    # 用于表头的填充色
    fill = PatternFill("solid", fgColor="FFF2CC")

    header = NamedStyle(name="header", font=header_font, fill=fill,
                        border=border_m, alignment=alignment)
    workbook.add_named_style(header)


def _register_content_style(workbook: Workbook):
    """注册content的风格"""
    # 字体
    content_font = Font(name="微软雅黑", size=10)
    # 边框
    line_t = Side(style="thin", color="000000")  # 细边框
    border_t = Border(top=line_t, bottom=line_t, left=line_t, right=line_t)

    # 对齐
    alignment = Alignment(horizontal="center", vertical="center")

    content = NamedStyle(name="content", font=content_font,
                         border=border_t, alignment=alignment)
    workbook.add_named_style(content)


def _register_styles(workbook: Workbook):
    """注册两种单元格风格"""
    if "header" not in workbook.style_names:
        _register_header_style(workbook)
    if "content" not in workbook.style_names:
        _register_content_style(workbook)


def _set_cell_styles(worksheet: Worksheet,
                     active_row: Optional[int] = None):
    """
    设置单元格的字体、边框等风格
    :param active_row: int型，从1开始
    """
    rows = worksheet.max_row
    columns = worksheet.max_column

    if active_row is not None:
        active_style = "header" if active_row < 3 else "content"
        for c in range(1, columns+1):
            col_letter = get_column_letter(c)
            worksheet[f"{col_letter}{active_row}"].style = active_style
    else:
        for r in range(1, rows+1):
            for c in range(1, columns+1):
                col_letter = get_column_letter(c)
                if r < 3:
                    worksheet[f"{col_letter}{r}"].style = "header"
                else:
                    worksheet[f"{col_letter}{r}"].style = "content"


def _move_merged_cells(merged_cells: List[MergedCellRange],
                       cur_max_col: int,
                       offset: int):
    """
    对部分merged cell向右移动，为新增的内容预留列
    :param merged_cells: 表示所有列单元格归一化的单元
    :param cur_max_col: int型，表示当前遍历到的单元格最大的列表
    :param offset: int型，表示插入列的数量
    :return:
    """
    for merged_cell in merged_cells:
        if merged_cell.min_col > cur_max_col:
            merged_cell.shift(offset, 0)


def _add_new_record(worksheet: Worksheet,
                    label_dict: Dict[str, Dict[str, Any]]):
    """向workbook中新增一条记录"""
    row_nums = worksheet.max_row
    next_row = row_nums + 1  # 表示插入的下一行
    # 除去前面两个行merged cell之外，后面几个列merged_cell
    merged_cell_ranges = worksheet.merged_cell_ranges
    # 只保留列merged_cell
    filter_merged_cell_ranges = [merged_cell for merged_cell in merged_cell_ranges
                                 if merged_cell.min_row == merged_cell.max_row]
    # 定义一个标识是否插入了新的列的标志位
    add_new_column_flag = False
    # 对于每一个大类（基本信息/画面信息/故事信息/角色信息/其他信息）
    for (header, info), merged_cell in zip(label_dict.items(), filter_merged_cell_ranges):
        # 获取当前merged_cell最小列和最大列的标号
        min_col, max_col = merged_cell.min_col, merged_cell.max_col
        # 检测当前header中列的数量和info的长度是否一致
        info_len = len(info)
        col_len = max_col - min_col + 1
        # 如果当前有多个信息，则需要插入列
        if info_len > col_len:
            add_new_column_flag = True
            _move_merged_cells(merged_cells=filter_merged_cell_ranges,
                               cur_max_col=max_col,
                               offset=info_len - col_len)
            # 将原来merged的部分unmerged
            worksheet.unmerge_cells(start_row=1, start_column=min_col,
                                    end_row=1, end_column=max_col)
            update_max_col = min_col + info_len - 1
            worksheet.merge_cells(start_row=1, start_column=min_col,
                                  end_row=1, end_column=update_max_col)
        for i, (k, v) in enumerate(info.items()):
            cur_column = min_col + i
            sub_header_value = worksheet.cell(2, column=cur_column).value
            if sub_header_value:
                if sub_header_value == k:
                    worksheet.cell(next_row, column=cur_column).value = v
                else:
                    logger.error(f"第{cur_column}列赋值时出现错误，当前列的名称为{sub_header_value}，"
                                 f"但是输入的键为{k}，输入的值为{v}。")
            else:
                worksheet.cell(2, column=cur_column).value = k
                worksheet.cell(next_row, column=cur_column).value = v
    # 调节行高和列宽
    if add_new_column_flag:
        _adjust_cell_width_and_height(worksheet,
                                      max_width=150,
                                      add_width=10,
                                      fixed_height=20)
        _set_cell_styles(worksheet)
    else:
        _adjust_cell_height(worksheet,
                            fixed_height=20,
                            active_row=next_row)
        _set_cell_styles(worksheet,
                         active_row=next_row)


def _new_table(worksheet: Worksheet,
               label_dict: Dict[str, Dict[str, Any]]):
    """新建一个表格"""
    # 注意worksheet的行和列标都是从1开始计数的
    # 在第1行第1列创建ID，第1行第2列创建"作品名称"
    worksheet.cell(row=1, column=1).value = "ID"
    worksheet.cell(row=1, column=2).value = "作品名称"
    # 合并第1列中第1行和第2行，以及第2列中的第1行和第2行
    worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

    # 读取数据中的每一段信息并保存
    # 由于前面两个字段占用了2列，所以接下来从第3列开始
    before_column, cur_column = 3, 3
    for header, info in label_dict.items():
        # 首先对第1行第1层header赋值
        worksheet.cell(1, cur_column).value = header
        for k, v in info.items():
            worksheet.cell(2, cur_column).value = k
            worksheet.cell(3, cur_column).value = v
            cur_column += 1
        # 合并第1行的before到cur列
        # 并更新before
        worksheet.merge_cells(start_row=1, start_column=before_column, end_row=1, end_column=cur_column - 1)
        before_column = cur_column

    # 调节行高和列宽
    _adjust_cell_width_and_height(worksheet,
                                  max_width=150,
                                  add_width=10,
                                  fixed_height=20)
    # 设置风格
    _set_cell_styles(worksheet)


def save_as_excel(to_file: str, novel_name: str, label: Label):
    """
    将标签保存到excel文件中
    :param to_file: str型，表示保存到的文件名
    :param novel_name: str型，小说的名称
    :param label: Label型，表示标签对象
    :return:
    """
    if not to_file.endswith(".xlsx"):
        filename_list = to_file.split(".")
        filename_list[-1] = "xlsx"
        to_file = ".".join(filename_list)

    # 将当前的标签转化为dict型
    label_dict = label.to_json()

    # 如果存在该文件则打开
    if os.path.isfile(to_file):
        workbook = load_workbook(to_file)
        # 获取当前第一个sheet作为活跃的sheet对象
        worksheet = workbook.active
        # 注册styles
        _register_styles(workbook)
        if worksheet.max_row > 2:
            _add_new_record(worksheet, label_dict)
        else:
            _new_table(worksheet, label_dict)

    else:
        # 创建一个excel对象
        workbook = Workbook()
        worksheet = workbook.active
        # 注册styles
        _register_styles(workbook)
        _new_table(worksheet, label_dict)

    # 设置id，以及novel_name
    max_row = worksheet.max_row
    worksheet.cell(row=max_row, column=1).value = (max_row - 2)
    worksheet.cell(row=max_row, column=2).value = novel_name
    # 保存文件
    workbook.save(to_file)
